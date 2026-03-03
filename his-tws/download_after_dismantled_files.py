import os
import ssl
import urllib.request

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "tws_sync_file-results.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "files-corrupt")
BASE_URL = "https://global.hino.co.id/his/protected/attachment/after_dismantled"


def read_filenames_from_excel(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {str(col).strip().lower(): idx for idx, col in enumerate(header_row)}

    if "filename" not in header_map:
        raise ValueError("Kolom 'filename' tidak ditemukan di file Excel.")

    filename_col_idx = header_map["filename"]

    filenames = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        value = row[filename_col_idx]
        if not value:
            continue
        value_str = str(value).strip()
        if value_str:
            filenames.append(value_str)

    return filenames


def download_file(filename, output_dir, base_url, context):
    url = f"{base_url}/{filename}"
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, filename)

    with urllib.request.urlopen(url, context=context) as response:
        content = response.read()

    with open(output_path, "wb") as f:
        f.write(content)

    return output_path


def main():
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"File Excel tidak ditemukan: {EXCEL_PATH}")

    filenames = read_filenames_from_excel(EXCEL_PATH)
    if not filenames:
        print("Tidak ada filename yang ditemukan di Excel.")
        return

    print(f"Total filename yang akan di-download: {len(filenames)}")

    context = ssl.create_default_context()
    success_count = 0

    for idx, filename in enumerate(filenames, start=1):
        try:
            print(f"[{idx}/{len(filenames)}] Download {filename} ...", end=" ")
            output_path = download_file(filename, OUTPUT_DIR, BASE_URL, context)
            success_count += 1
            print(f"OK -> {output_path}")
            if success_count >= 100:
                print(f"Berhasil download {success_count} file, proses dihentikan.")
                break
        except Exception as e:
            print(f"GAGAL ({e})")


if __name__ == "__main__":
    main()
