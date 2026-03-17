import json
import re
import sys
import time
import warnings
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl
import requests


warnings.filterwarnings("ignore")


DEFAULT_BASEURL = "https://hdcs.hinodms.co.id"
USERINFO = [{"LoginID": "DMSAPI001", "Password": "password.123"}]
TIMEOUT_SECONDS = 10
SLEEP_SECONDS = 0.05
MAX_CELL_CHARS = 1800


def _clean_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _safe_cell(text: Any) -> str:
    if text is None:
        return ""
    t = str(text)
    if len(t) > MAX_CELL_CHARS:
        return t[:MAX_CELL_CHARS] + "..."
    return t


def _flatten_postman_items(items: List[Dict[str, Any]], parents: List[str]) -> Iterable[Tuple[List[str], Dict[str, Any]]]:
    for it in items or []:
        name = it.get("name", "")
        if "item" in it:
            yield from _flatten_postman_items(it.get("item") or [], parents + [name])
            continue
        req = it.get("request")
        if req:
            yield parents + [name], req


def _resolve_url(req: Dict[str, Any]) -> str:
    url = req.get("url") or {}
    raw = (url.get("raw") or "").replace("{{baseURL}}", DEFAULT_BASEURL)
    if raw.startswith("http://") or raw.startswith("https://"):
        return raw

    protocol = url.get("protocol") or "https"
    host = url.get("host") or ""
    if isinstance(host, list):
        host_str = ".".join([str(x) for x in host if x is not None])
    else:
        host_str = str(host)
    host_str = host_str.replace("{{baseURL}}", DEFAULT_BASEURL)

    if host_str.startswith("http://") or host_str.startswith("https://"):
        base = host_str
    else:
        base = f"{protocol}://{host_str}"

    path = url.get("path") or ""
    if isinstance(path, list):
        p = "/".join([str(x) for x in path if x is not None])
    else:
        p = str(path)
    if not p.startswith("/"):
        p = "/" + p

    return base + p


def _short_path_from_url(url: str) -> str:
    m = re.search(r"/index\.php/(.+)$", url)
    if m:
        tail = m.group(1)
        parts = [p for p in tail.split("/") if p]
        if len(parts) >= 2:
            return "/" + parts[-2] + "/" + parts[-1]
        if len(parts) == 1:
            return "/" + parts[0]

    parts = [p for p in url.split("/") if p]
    if len(parts) >= 2:
        return "/" + parts[-2] + "/" + parts[-1]
    return ""


def _headers_from_postman(req: Dict[str, Any]) -> Dict[str, str]:
    headers: Dict[str, str] = {}
    for h in req.get("header") or []:
        k = h.get("key")
        if not k:
            continue
        headers[str(k)] = "" if h.get("value") is None else str(h.get("value"))
    if "Content-Type" not in headers:
        headers["Content-Type"] = "application/json"
    return headers


def _normalized_request_body(req: Dict[str, Any]) -> str:
    body = req.get("body") or {}
    if body.get("mode") != "raw":
        return ""
    raw = (body.get("raw") or "").replace("\r\n", "\n")

    lines: List[str] = []
    for line in raw.split("\n"):
        if "//" in line:
            idx = line.find("//")
            prefix = line[:idx]
            if "http://" not in prefix and "https://" not in prefix:
                line = prefix
        if line.strip() == "":
            continue
        lines.append(line)
    raw = "\n".join(lines)
    raw = re.sub(r",\s*([}\]])", r"\\1", raw)

    try:
        data = json.loads(raw)
        if isinstance(data, dict):
            data["UserInfo"] = USERINFO
        return json.dumps(data, ensure_ascii=False)
    except Exception:
        return raw


def _sample_response(text: Any) -> str:
    if text is None:
        return ""
    t = str(text)
    try:
        data = json.loads(t)
        if isinstance(data, list):
            return json.dumps(data[:3], ensure_ascii=False)
        if isinstance(data, dict):
            for key in ["data", "result", "items", "rows"]:
                if key in data and isinstance(data[key], list):
                    data2 = dict(data)
                    data2[key] = data[key][:3]
                    return json.dumps(data2, ensure_ascii=False)
            keys = list(data.keys())[:15]
            data2 = {k: data[k] for k in keys}
            return json.dumps(data2, ensure_ascii=False)
    except Exception:
        pass
    return t


def _ensure_extra_columns(ws) -> Tuple[int, int, int]:
    wanted = ["Postman Name", "Request Body", "Response Body"]
    seen: Dict[str, List[int]] = {k: [] for k in wanted}

    last_non_empty = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None or str(v).strip() == "":
            continue
        last_non_empty = c
        key = str(v).strip()
        if key in seen:
            seen[key].append(c)

    first: Dict[str, int] = {}
    for k in wanted:
        if seen[k]:
            first[k] = seen[k][0]

    if not first:
        start = last_non_empty + 1
        for i, k in enumerate(wanted):
            ws.cell(1, start + i).value = k
            first[k] = start + i
        return first["Postman Name"], first["Request Body"], first["Response Body"]

    max_existing = max(first.values())
    for k in wanted:
        if k not in first:
            max_existing += 1
            ws.cell(1, max_existing).value = k
            first[k] = max_existing

    for k in wanted:
        dest = first[k]
        for src in seen[k][1:]:
            for r in range(2, ws.max_row + 1):
                if ws.cell(r, dest).value is None and ws.cell(r, src).value is not None:
                    ws.cell(r, dest).value = ws.cell(r, src).value
            ws.cell(1, src).value = None
            for r in range(2, ws.max_row + 1):
                ws.cell(r, src).value = None

    return first["Postman Name"], first["Request Body"], first["Response Body"]


def _find_row_by_short_path(row_texts: List[Tuple[int, str]], short_path: str) -> Optional[int]:
    sp = short_path.lower()
    if not sp:
        return None
    for r, txt in row_texts:
        if sp in txt:
            return r
    return None


def main() -> int:
    if "--headers-only" in sys.argv:
        base_dir = Path(__file__).resolve().parent
        xlsx_path = base_dir / "List All API HDCS Server.xlsx"
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb[wb.sheetnames[0]]
        col_postman, col_req_body, col_resp_body = _ensure_extra_columns(ws)
        wb.save(str(xlsx_path))
        print("headers_ready", col_postman, col_req_body, col_resp_body)
        return 0

    base_dir = Path(__file__).resolve().parent
    xlsx_path = base_dir / "List All API HDCS Server.xlsx"
    collections = [
        base_dir / "API HMSI for Dealer.postman_collection.json",
        base_dir / "Digital Control Board.postman_collection.json",
    ]

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb[wb.sheetnames[0]]

    col_postman, col_req_body, col_resp_body = _ensure_extra_columns(ws)

    col_no = 1
    col_integration_name = 2
    col_source_system = 4
    col_integrator_db = 5
    col_target_system = 6
    col_action_desc = 7
    col_integration_type = 12
    col_method = 13
    col_status = 16
    col_testing_result = 18

    row_texts: List[Tuple[int, str]] = []
    max_no = 0
    for r in range(2, ws.max_row + 1):
        action_desc = ws.cell(r, col_action_desc).value
        row_texts.append((r, _clean_str(action_desc).lower()))
        try:
            v = ws.cell(r, col_no).value
            if v is not None:
                max_no = max(max_no, int(v))
        except Exception:
            pass

    processed = 0
    matched = 0
    added = 0
    success = 0
    failed = 0

    session = requests.Session()

    for coll_path in collections:
        coll = json.loads(coll_path.read_text(encoding="utf-8"))
        coll_name = coll.get("info", {}).get("name") or coll_path.name
        for name_parts, req in _flatten_postman_items(coll.get("item") or [], [coll_name]):
            processed += 1

            postman_name = " / ".join([p for p in name_parts if p])
            method = (req.get("method") or "GET").upper()
            url = _resolve_url(req)
            short_path = _short_path_from_url(url)
            body = _normalized_request_body(req)
            headers = _headers_from_postman(req)

            row = _find_row_by_short_path(row_texts, short_path)
            if row is None:
                ws.append([None] * max(ws.max_column, col_resp_body))
                row = ws.max_row
                max_no += 1
                ws.cell(row, col_no).value = max_no
                ws.cell(row, col_integration_name).value = postman_name.split(" / ")[-1]
                ws.cell(row, col_source_system).value = "HDCS System"
                ws.cell(row, col_integrator_db).value = "NEWHINODMS (SQL Server)"
                ws.cell(row, col_target_system).value = "Hino DMS API (api.hinodms.co.id)"
                ws.cell(row, col_action_desc).value = short_path or url
                ws.cell(row, col_integration_type).value = "REST API"
                ws.cell(row, col_method).value = method
                ws.cell(row, col_status).value = "Active"
                row_texts.append((row, _clean_str(ws.cell(row, col_action_desc).value).lower()))
                added += 1
            else:
                matched += 1

            ws.cell(row, col_postman).value = postman_name

            status_code: Optional[int] = None
            resp_text = ""
            err = ""
            try:
                resp = session.request(
                    method,
                    url,
                    headers=headers,
                    data=body if body else None,
                    timeout=TIMEOUT_SECONDS,
                    verify=False,
                )
                status_code = resp.status_code
                resp_text = resp.text
            except Exception as e:
                err = str(e)

            ok = status_code is not None and 200 <= status_code < 300
            if ok:
                success += 1
                testing_val = f"SUCCESS ({status_code})"
            else:
                failed += 1
                testing_val = f"FAILED ({status_code if status_code is not None else err})"

            ws.cell(row, col_testing_result).value = testing_val
            ws.cell(row, col_req_body).value = _safe_cell(body)
            ws.cell(row, col_resp_body).value = _safe_cell(_sample_response(resp_text))

            time.sleep(SLEEP_SECONDS)

    wb.save(str(xlsx_path))

    print("DONE")
    print("processed", processed)
    print("matched", matched)
    print("added", added)
    print("success", success)
    print("failed", failed)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
